import os
import re
import uuid
import copy
import subprocess
from datetime import datetime, date
from flask import Flask, request, jsonify, render_template, send_file
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.datetime import from_excel as _from_excel

_DATE_NF_RE = re.compile(r'[yYdD]|(?<!\[)[mM](?!\])', re.ASCII)

app = Flask(__name__)
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

workbooks = {}

# openpyxl indexed color table (Excel標準56色)
INDEXED_COLORS = [
    '000000','FFFFFF','FF0000','00FF00','0000FF','FFFF00','FF00FF','00FFFF',
    '000000','FFFFFF','FF0000','00FF00','0000FF','FFFF00','FF00FF','00FFFF',
    '800000','008000','000080','808000','800080','008080','C0C0C0','808080',
    '9999FF','993366','FFFFCC','CCFFFF','660066','FF8080','0066CC','CCCCFF',
    '000080','FF00FF','FFFF00','00FFFF','800080','800000','008080','0000FF',
    '00CCFF','CCFFFF','CCFFCC','FFFF99','99CCFF','FF99CC','CC99FF','FFCC99',
    '3366FF','33CCCC','99CC00','FFCC00','FF9900','FF6600','666699','969696',
    '003366','339966','003300','333300','993300','993366','333399','333333',
]


# ─────────────────────────────────────────────
#  CORS（Capacitorアプリからの接続を許可）
# ─────────────────────────────────────────────
@app.after_request
def add_cors(response):
    response.headers['Access-Control-Allow-Origin']  = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response

@app.route('/api/<path:path>', methods=['OPTIONS'])
def options_handler(path):
    return '', 204


# ─────────────────────────────────────────────
#  ユーティリティ
# ─────────────────────────────────────────────
def fmt(v):
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y/%m/%d')
    return str(v)


def _resolve_date(cell, dc, v):
    """シリアル日付数値を datetime へ変換。変換できない場合は Render ログに出力。"""
    import sys
    if not isinstance(v, (int, float)):
        return v
    # ① data_only=False 側が datetime を返していればそれを直接使う
    raw = cell.value
    if isinstance(raw, (datetime, date)):
        return raw
    # ② is_date (openpyxl 判定)
    for c in (cell, dc):
        if c.is_date:
            try:
                return _from_excel(v)
            except Exception:
                return v
    # ③ 書式文字列に日付トークン (y/d/m)
    for c in (cell, dc):
        nf = c.number_format or ''
        if nf and nf not in ('General', '@', '') and _DATE_NF_RE.search(nf):
            try:
                return _from_excel(v)
            except Exception:
                return v
    # ④ 変換できなかった場合: 診断ログ出力（Render の Logs で確認可）
    if isinstance(v, (int, float)) and isinstance(raw, (int, float)):
        print(f'[date_debug] v={v!r} raw={raw!r} '
              f'cell.nf={cell.number_format!r} dc.nf={dc.number_format!r} '
              f'cell.is_date={cell.is_date} dc.is_date={dc.is_date}',
              file=sys.stderr, flush=True)
    return v


def get_cell_bg(cell):
    """セル背景色を '#RRGGBB' で返す。色なしは ''"""
    try:
        fill = cell.fill
        if not fill or fill.fill_type != 'solid':
            return ''
        color = fill.fgColor
        if color.type == 'rgb':
            rgb = color.rgb  # 'AARRGGBB' (8桁) or 'RRGGBB' (6桁)
            if len(rgb) == 8:
                if rgb[:2].upper() == '00':
                    return ''
                rgb = rgb[2:]
            return '#' + rgb
        elif color.type == 'indexed':
            idx = color.indexed
            if idx == 64:
                return ''
            if 0 <= idx < len(INDEXED_COLORS):
                return '#' + INDEXED_COLORS[idx]
    except Exception:
        pass
    return ''


def build_state(wb, wb_data):
    display  = {}
    formulas = {}
    bg       = {}
    for name in wb.sheetnames:
        ws   = wb[name]
        ws_d = wb_data[name]
        max_r = ws.max_row or 1
        max_c = ws.max_column or 1
        drows, frows, bgrows = [], [], []
        for r in range(1, max_r + 1):
            drow, frow, bgrow = [], [], []
            for c in range(1, max_c + 1):
                cell = ws.cell(r, c)
                dc   = ws_d.cell(r, c)
                raw  = cell.value

                if isinstance(raw, str) and raw.startswith('='):
                    frow.append(raw)
                else:
                    frow.append(fmt(raw))

                disp = dc.value
                if disp is not None:
                    disp = _resolve_date(cell, dc, disp)
                    drow.append(fmt(disp))
                elif isinstance(raw, str) and raw.startswith('='):
                    drow.append('')
                else:
                    drow.append(fmt(raw))

                bgrow.append(get_cell_bg(cell))

            drows.append(drow)
            frows.append(frow)
            bgrows.append(bgrow)
        display[name]  = drows
        formulas[name] = frows
        bg[name]       = bgrows
    return display, formulas, bg


def ensure_size(entry, sheet, ri, ci):
    for key in ('display', 'formulas', 'bg'):
        mat = entry[key][sheet]
        while len(mat) <= ri:
            cols = len(mat[0]) if mat else (ci + 1)
            mat.append(['' for _ in range(cols)])
        for row in mat:
            while len(row) <= ci:
                row.append('')


def state_response(entry):
    return jsonify({
        'success': True,
        'sheets':   list(entry['wb'].sheetnames),
        'values':   entry['display'],
        'formulas': entry['formulas'],
        'bg':       entry['bg'],
    })


# ─────────────────────────────────────────────
#  ルート
# ─────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')


def _is_local():
    return request.remote_addr in ('127.0.0.1', '::1')


def _osascript(script):
    r = subprocess.run(['osascript', '-e', script], capture_output=True, text=True)
    return r.returncode, r.stdout.strip()


@app.route('/api/open_dialog', methods=['POST'])
def open_dialog():
    """Mac本体からのアクセス時のみ: ネイティブダイアログでファイルを選択して開く"""
    if not _is_local():
        return jsonify({'error': 'local_only'}), 403

    code, path = _osascript(
        'tell application "Finder" to activate\n'
        'set f to (choose file with prompt "Excelファイルを選択"'
        ' of type {"xlsx","xls","org.openxmlformats.spreadsheetml.sheet"})\n'
        'return POSIX path of f'
    )
    if code != 0 or not path:
        return jsonify({'cancelled': True})
    if not os.path.exists(path):
        return jsonify({'error': 'ファイルが見つかりません'}), 400

    file_id  = str(uuid.uuid4())
    filename = os.path.basename(path)
    wb   = openpyxl.load_workbook(path, data_only=False)
    wb_d = openpyxl.load_workbook(path, data_only=True)
    display, formulas, bg = build_state(wb, wb_d)
    workbooks[file_id] = {
        'wb': wb, 'filepath': path, 'filename': filename,
        'display': display, 'formulas': formulas, 'bg': bg,
    }
    return jsonify({'file_id': file_id, 'filename': filename,
                    'sheets': list(wb.sheetnames),
                    'values': display, 'formulas': formulas, 'bg': bg})


@app.route('/api/save_local/<file_id>', methods=['POST'])
def save_local(file_id):
    """元のパスに上書き保存（Macダイアログで開いたファイル用）"""
    if file_id not in workbooks:
        return jsonify({'error': 'Not found'}), 404
    entry = workbooks[file_id]
    entry['wb'].save(entry['filepath'])
    return jsonify({'success': True, 'path': entry['filepath']})


@app.route('/api/save_dialog/<file_id>', methods=['POST'])
def save_dialog(file_id):
    """Mac上で保存先ダイアログを開き保存する"""
    if file_id not in workbooks:
        return jsonify({'error': 'Not found'}), 404
    entry = workbooks[file_id]
    fname = entry['filename'].replace('"', '')

    code, path = _osascript(
        f'tell application "Finder" to activate\n'
        f'set f to (choose file name with prompt "保存先を選択してください"'
        f' default name "{fname}")\n'
        f'return POSIX path of f'
    )
    if code != 0 or not path:
        return jsonify({'cancelled': True})

    if not path.endswith('.xlsx'):
        path += '.xlsx'

    entry['wb'].save(path)
    entry['filepath'] = path
    return jsonify({'success': True, 'path': os.path.basename(path)})


@app.route('/api/upload', methods=['POST'])
def upload():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'ファイルがありません'}), 400
    file_id  = str(uuid.uuid4())
    filepath = os.path.join(UPLOAD_FOLDER, f'{file_id}.xlsx')
    f.save(filepath)
    wb   = openpyxl.load_workbook(filepath, data_only=False)
    wb_d = openpyxl.load_workbook(filepath, data_only=True)
    display, formulas, bg = build_state(wb, wb_d)
    workbooks[file_id] = {
        'wb': wb, 'filepath': filepath, 'filename': f.filename,
        'display': display, 'formulas': formulas, 'bg': bg,
    }
    return jsonify({'file_id': file_id, 'filename': f.filename,
                    'sheets': list(wb.sheetnames),
                    'values': display, 'formulas': formulas, 'bg': bg})


@app.route('/api/new', methods=['POST'])
def new_file():
    file_id  = str(uuid.uuid4())
    filepath = os.path.join(UPLOAD_FOLDER, f'{file_id}.xlsx')
    wb = openpyxl.Workbook()
    wb.active.title = 'Sheet1'
    wb.save(filepath)
    empty = [['' for _ in range(8)] for _ in range(20)]
    workbooks[file_id] = {
        'wb': wb, 'filepath': filepath, 'filename': '新しいファイル.xlsx',
        'display':  copy.deepcopy(empty),
        'formulas': copy.deepcopy(empty),
        'bg':       copy.deepcopy(empty),
    }
    for key in ('display', 'formulas', 'bg'):
        workbooks[file_id][key] = {'Sheet1': workbooks[file_id][key]}
    return jsonify({'file_id': file_id, 'filename': '新しいファイル.xlsx',
                    'sheets': ['Sheet1'],
                    'values': workbooks[file_id]['display'],
                    'formulas': workbooks[file_id]['formulas'],
                    'bg': workbooks[file_id]['bg']})


@app.route('/api/cell', methods=['POST'])
def update_cell():
    d = request.json
    file_id = d['file_id']
    if file_id not in workbooks:
        return jsonify({'error': 'Not found'}), 404
    entry  = workbooks[file_id]
    sheet  = d['sheet']
    ri, ci = int(d['row']), int(d['col'])
    value  = d['value']

    ws   = entry['wb'][sheet]
    cell = ws.cell(ri + 1, ci + 1)
    if value == '':
        cell.value = None
    elif isinstance(value, str) and value.startswith('='):
        cell.value = value
    else:
        try:
            cell.value = int(value)
        except ValueError:
            try:
                cell.value = float(value)
            except ValueError:
                cell.value = value

    entry['wb'].save(entry['filepath'])
    ensure_size(entry, sheet, ri, ci)

    if isinstance(value, str) and value.startswith('='):
        entry['formulas'][sheet][ri][ci] = value
    else:
        entry['formulas'][sheet][ri][ci] = value
        entry['display'][sheet][ri][ci]  = value

    return state_response(entry)


@app.route('/api/color', methods=['POST'])
def update_color():
    d = request.json
    file_id = d['file_id']
    if file_id not in workbooks:
        return jsonify({'error': 'Not found'}), 404
    entry  = workbooks[file_id]
    sheet  = d['sheet']
    ri, ci = int(d['row']), int(d['col'])
    color  = d.get('color', '')

    ws   = entry['wb'][sheet]
    cell = ws.cell(ri + 1, ci + 1)

    if color:
        hex_color = color.lstrip('#').upper()
        cell.fill = PatternFill(fill_type='solid', fgColor=hex_color, bgColor=hex_color)
    else:
        cell.fill = PatternFill(fill_type=None)

    entry['wb'].save(entry['filepath'])
    ensure_size(entry, sheet, ri, ci)
    entry['bg'][sheet][ri][ci] = color

    return state_response(entry)


@app.route('/api/add_row', methods=['POST'])
def add_row():
    d = request.json
    entry = workbooks[d['file_id']]
    sheet = d['sheet']
    ws    = entry['wb'][sheet]
    ws.append([None] * (ws.max_column or 1))
    entry['wb'].save(entry['filepath'])
    cols = len(entry['display'][sheet][0]) if entry['display'][sheet] else 1
    for key in ('display', 'formulas', 'bg'):
        entry[key][sheet].append(['' for _ in range(cols)])
    return state_response(entry)


@app.route('/api/add_col', methods=['POST'])
def add_col():
    d = request.json
    entry = workbooks[d['file_id']]
    sheet = d['sheet']
    ws    = entry['wb'][sheet]
    new_c = (ws.max_column or 0) + 1
    for r in range(1, (ws.max_row or 1) + 1):
        ws.cell(r, new_c).value = None
    entry['wb'].save(entry['filepath'])
    for key in ('display', 'formulas', 'bg'):
        for row in entry[key][sheet]:
            row.append('')
    return state_response(entry)


@app.route('/api/add_sheet', methods=['POST'])
def add_sheet():
    d = request.json
    entry = workbooks[d['file_id']]
    name  = d['name']
    if name in entry['wb'].sheetnames:
        return jsonify({'error': '同名のシートが既に存在します'}), 400
    entry['wb'].create_sheet(name)
    entry['wb'].save(entry['filepath'])
    empty = [['' for _ in range(8)] for _ in range(20)]
    for key in ('display', 'formulas', 'bg'):
        entry[key][name] = copy.deepcopy(empty)
    return state_response(entry)


@app.route('/api/save/<file_id>')
def save_file(file_id):
    if file_id not in workbooks:
        return jsonify({'error': 'Not found'}), 404
    entry = workbooks[file_id]
    fname = entry['filename']
    if not fname.endswith('.xlsx'):
        fname += '.xlsx'
    return send_file(entry['filepath'], as_attachment=True, download_name=fname)


if __name__ == '__main__':
    import socket
    try:
        local_ip = socket.gethostbyname(socket.gethostname())
    except Exception:
        local_ip = '不明'
    print(f'\n📊 Excel Editor 起動中')
    print(f'   Mac から:                http://localhost:8080')
    print(f'   iPhone から (同一Wi-Fi): http://{local_ip}:8080\n')
    app.run(host='0.0.0.0', port=8080, debug=False, threaded=True)
